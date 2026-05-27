"""文件解析工具 — 从上传文件中提取文本。

纯函数，无副作用，从 app/api/chat.py 提取而来。
"""

from __future__ import annotations

import re
from io import BytesIO
from pathlib import Path

import pdfplumber
import xlrd
from docx import Document
from openpyxl import load_workbook

MAX_FILE_TEXT_CHARS = 50000
ALLOWED_EXTS = {".txt", ".md", ".pdf", ".xlsx", ".xls", ".docx", ".doc"}


def normalize_name(name: str | None) -> str:
    raw = (name or "uploaded_file").strip()
    safe = re.sub(r"[^\w\-.一-鿿]", "_", raw)
    return safe[:120] or "uploaded_file"


def clip_text(text: str, max_chars: int = MAX_FILE_TEXT_CHARS) -> str:
    """截断文件文本，使用统一的 truncate_text 工具。"""
    from ...agent.context.token_utils import truncate_text
    return truncate_text(text, max_chars, mode="chars", marker="...[文件内容已截断]")


def _extract_pdf_text(content: bytes) -> str:
    parts: list[str] = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            txt = page.extract_text() or ""
            if txt.strip():
                parts.append(f"## Page {i}\n{txt.strip()}")
    return "\n\n".join(parts)


def _extract_xlsx_text(content: bytes) -> str:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sections: list[str] = []
    for ws in wb.worksheets:
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            values = ["" if v is None else str(v) for v in row]
            line = "\t".join(values).strip()
            if line:
                rows.append(line)
        if rows:
            sections.append(f"## Sheet: {ws.title}\n" + "\n".join(rows))
    return "\n\n".join(sections)


def _extract_xls_text(content: bytes) -> str:
    wb = xlrd.open_workbook(file_contents=content)
    sections: list[str] = []
    for sheet in wb.sheets():
        rows: list[str] = []
        for r in range(sheet.nrows):
            values = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            line = "\t".join(values).strip()
            if line:
                rows.append(line)
        if rows:
            sections.append(f"## Sheet: {sheet.name}\n" + "\n".join(rows))
    return "\n\n".join(sections)


def _extract_docx_text(content: bytes) -> str:
    doc = Document(BytesIO(content))
    parts = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    return "\n".join(parts)


def _extract_doc_text(content: bytes) -> str:
    for enc in ("utf-8", "gb18030", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法解析 .doc 文件内容")


def extract_file_text(filename: str, ext: str, content: bytes) -> str:
    """根据扩展名选择合适的解析器提取文本。"""
    if ext in {".txt", ".md"}:
        for enc in ("utf-8", "gb18030", "latin-1"):
            try:
                return content.decode(enc)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"文件 {filename} 编码无法识别")
    if ext == ".pdf":
        return _extract_pdf_text(content)
    if ext == ".xlsx":
        return _extract_xlsx_text(content)
    if ext == ".xls":
        return _extract_xls_text(content)
    if ext == ".docx":
        return _extract_docx_text(content)
    if ext == ".doc":
        return _extract_doc_text(content)
    raise ValueError(f"不支持的文件类型: {ext}")


def build_user_message(
    message: str,
    filename: str | None = None,
    file_text: str | None = None,
) -> str:
    """构建最终的用户消息（可能包含上传文件内容）。"""
    text = (message or "").strip()
    if not filename or not file_text:
        return text
    clipped = clip_text(file_text.strip())
    file_block = f"[用户上传文件]\n文件名: {filename}\n\n{clipped}\n[/用户上传文件]"
    return file_block if not text else f"{file_block}\n\n用户问题：{text}"
