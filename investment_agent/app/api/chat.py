import json
import re
import uuid
from io import BytesIO
from pathlib import Path

import pdfplumber
import xlrd
from docx import Document
from fastapi import APIRouter, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from pydantic import BaseModel

from ...agent import AgentRunner
from ..config_factory import load_agent_run_config
from ..storage import SqliteStorage

router = APIRouter(prefix="/api/chat", tags=["chat"])


class ChatRequest(BaseModel):
    session_id: str | None = None  # 为空则创建新会话
    message: str
    agent_id: str | None = None   # 为空则使用默认 Agent


# 文件上传限制
ALLOWED_EXTS = {".txt", ".md", ".pdf", ".xlsx", ".xls", ".docx", ".doc"}
MAX_UPLOAD_BYTES = 10 * 1024 * 1024  # 10MB
MAX_FILE_TEXT_CHARS = 50000  # 提取文本上限


def _normalize_name(name: str | None) -> str:
    raw = (name or "uploaded_file").strip()
    safe = re.sub(r"[^\w\-.一-鿿]", "_", raw)
    return safe[:120] or "uploaded_file"


def _clip_text(text: str, max_chars: int = MAX_FILE_TEXT_CHARS) -> str:
    if len(text) <= max_chars:
        return text
    return text[:max_chars] + "\n...[文件内容已截断]"


def _extract_pdf_text(content: bytes) -> str:
    parts: list[str] = []
    with pdfplumber.open(BytesIO(content)) as pdf:
        for i, page in enumerate(pdf.pages, start=1):
            txt = page.extract_text() or ""
            if txt.strip():
                parts.append(f"## Page {i}\n{txt.strip()}")
    return "\n\n".join(parts)


def _extract_xlsx_text(content: bytes) -> str:
    wb = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
    sections: list[str] = []
    for ws in wb.worksheets:
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            values = ["" if v is None else str(v) for v in row]
            line = "\t".join(values).strip()
            if line:
                rows.append(line)
        if rows:
            sections.append(f"## Sheet: {ws.title}\n" + "\n".join(rows))
    return "\n\n".join(sections)


def _extract_xls_text(content: bytes) -> str:
    wb = xlrd.open_workbook(file_contents=content)
    sections: list[str] = []
    for sheet in wb.sheets():
        rows: list[str] = []
        for r in range(sheet.nrows):
            values = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            line = "\t".join(values).strip()
            if line:
                rows.append(line)
        if rows:
            sections.append(f"## Sheet: {sheet.name}\n" + "\n".join(rows))
    return "\n\n".join(sections)


def _extract_docx_text(content: bytes) -> str:
    doc = Document(BytesIO(content))
    parts = [p.text.strip() for p in doc.paragraphs if p.text and p.text.strip()]
    return "\n".join(parts)


def _extract_doc_text(content: bytes) -> str:
    for enc in ("utf-8", "gb18030", "latin-1"):
        try:
            return content.decode(enc)
        except UnicodeDecodeError:
            continue
    raise ValueError("无法解析 .doc 文件内容")


def _extract_file_text(filename: str, ext: str, content: bytes) -> str:
    if ext in {".txt", ".md"}:
        for enc in ("utf-8", "gb18030", "latin-1"):
            try:
                return content.decode(enc)
            except UnicodeDecodeError:
                continue
        raise ValueError(f"文件 {filename} 编码无法识别")
    if ext == ".pdf":
        return _extract_pdf_text(content)
    if ext == ".xlsx":
        return _extract_xlsx_text(content)
    if ext == ".xls":
        return _extract_xls_text(content)
    if ext == ".docx":
        return _extract_docx_text(content)
    if ext == ".doc":
        return _extract_doc_text(content)
    raise ValueError(f"不支持的文件类型: {ext}")


def _build_user_message(message: str, filename: str | None, file_text: str | None) -> str:
    text = (message or "").strip()
    if not filename or not file_text:
        return text
    clipped = _clip_text(file_text.strip())
    file_block = f"[用户上传文件]\n文件名: {filename}\n\n{clipped}\n[/用户上传文件]"
    return file_block if not text else f"{file_block}\n\n用户问题：{text}"


async def _parse_request_payload(request: Request) -> tuple[str | None, str, str | None, str | None, str | None]:
    """解析请求体：支持 JSON 和 multipart/form-data（带文件上传）两种格式"""
    content_type = request.headers.get("content-type", "")

    if "application/json" in content_type:
        body = ChatRequest.model_validate(await request.json())
        return body.session_id, body.message, body.agent_id, None, None

    if "multipart/form-data" in content_type:
        form = await request.form()
        session_id = (form.get("session_id") or "").strip() or None
        message = str(form.get("message") or "").strip()
        agent_id = (form.get("agent_id") or "").strip() or None
        upload = form.get("file")

        filename_attr = getattr(upload, "filename", None)
        if not upload or not filename_attr:
            return session_id, message, agent_id, None, None

        filename = _normalize_name(str(filename_attr))
        ext = Path(filename).suffix.lower()
        if ext not in ALLOWED_EXTS:
            raise HTTPException(status_code=400, detail=f"不支持的文件类型: {ext or 'unknown'}")

        content = await upload.read()
        if len(content) > MAX_UPLOAD_BYTES:
            raise HTTPException(status_code=400, detail="文件过大，最大支持 10MB")

        try:
            file_text = _extract_file_text(filename, ext, content)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"文件解析失败: {e}")

        if not file_text.strip():
            raise HTTPException(status_code=400, detail="文件内容为空或不可提取")

        return session_id, message, agent_id, filename, file_text

    raise HTTPException(status_code=415, detail="仅支持 application/json 或 multipart/form-data")


@router.post("")
async def start_chat(request: Request):
    """发起对话：解析请求，加载配置，委托 AgentRunner 创建引擎"""
    session_id, message, agent_id, filename, file_text = await _parse_request_payload(request)
    final_message = _build_user_message(message, filename, file_text)
    if not final_message.strip():
        raise HTTPException(status_code=400, detail="消息和文件不能同时为空")

    config = await load_agent_run_config(agent_id)
    runner = AgentRunner(storage=SqliteStorage())
    task_id, session_id = await runner.start(
        session_id=session_id or str(uuid.uuid4()),
        config=config,
        message=final_message,
    )
    return {"task_id": task_id, "session_id": session_id}


@router.get("/{task_id}/stream")
async def stream_chat(task_id: str):
    """SSE 流式端点：委托 AgentRunner 执行引擎，yield SSE 事件"""
    runner = AgentRunner(storage=SqliteStorage())

    async def event_stream():
        try:
            async for event in runner.run(task_id):
                yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
            await runner.save_response()
        finally:
            runner.cleanup(task_id)

    return StreamingResponse(
        event_stream(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


@router.post("/{task_id}/interrupt")
async def interrupt_chat(task_id: str):
    ok = AgentRunner.interrupt(task_id)
    return {"success": ok}
